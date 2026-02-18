"""
Enhanced XLSX Processor.
Uses openpyxl directly for merged cell support, comments, and data type preservation.
"""

import io
import logging
from typing import List, Dict, Any, Optional

from processors.base import FileProcessor, TextExtractionResult

logger = logging.getLogger(__name__)


class XLSXProcessor(FileProcessor):
    """Process XLSX/XLS files with merged cell handling and comment extraction."""

    def can_process(self, mime_type: str, extension: str) -> bool:
        return (mime_type in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'
        ] or extension.lower() in ['.xlsx', '.xls'])

    def extract_text(self, file_bytes: bytes, filename: str, file_id: str) -> TextExtractionResult:
        import openpyxl

        result = TextExtractionResult(file_id, filename, "xlsx")
        result.extraction_method = "openpyxl-enhanced"

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)

            all_comments = {}

            for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]

                # Build merged cell map — fill merged regions with the top-left value
                merged_map = self._build_merged_cell_map(ws)

                # Extract data with merged cell support
                rows_data = []
                max_col = ws.max_column or 0
                max_row = ws.max_row or 0

                if max_row == 0 or max_col == 0:
                    result.add_page(sheet_num, f"Sheet: {sheet_name}\n(empty)", "openpyxl-enhanced",
                                    sheet_name=sheet_name, rows=0, columns=0)
                    continue

                for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    row_text = []
                    for cell in row:
                        # Check merged map first
                        val = merged_map.get((cell.row, cell.column))
                        if val is None:
                            val = cell.value

                        # Format value based on type
                        row_text.append(self._format_cell_value(val, cell))
                    rows_data.append(row_text)

                # Format sheet as text
                sheet_text = f"Sheet: {sheet_name}\n\n"

                if rows_data:
                    # Calculate column widths for alignment
                    col_widths = self._calculate_column_widths(rows_data)
                    for row in rows_data:
                        formatted_cells = []
                        for i, cell_val in enumerate(row):
                            width = col_widths[i] if i < len(col_widths) else 10
                            formatted_cells.append(cell_val.ljust(width))
                        sheet_text += " | ".join(formatted_cells) + "\n"

                result.add_page(sheet_num, sheet_text, "openpyxl-enhanced",
                                sheet_name=sheet_name,
                                rows=len(rows_data),
                                columns=max_col)

                # Add as structured table
                if rows_data:
                    result.add_table(sheet_num, rows_data, 0)

                # Extract comments
                comments = self._extract_comments(ws)
                if comments:
                    all_comments[sheet_name] = comments

            wb.close()

            result.quality_metrics.average_confidence = 100.0
            result.quality_metrics.completeness_score = min(
                100, result.quality_metrics.total_text_length / 100
            )

            result.metadata = {
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "has_comments": bool(all_comments)
            }
            if all_comments:
                result.metadata["comments"] = all_comments

            logger.info(
                f"Extracted text from XLSX: {filename} "
                f"({len(wb.sheetnames)} sheets, {result.quality_metrics.total_tables} tables)"
            )

        except Exception as e:
            logger.error(f"Failed to process XLSX {filename}: {e}")
            # Fallback to pandas for basic extraction
            try:
                return self._extract_with_pandas(file_bytes, filename, file_id)
            except Exception:
                raise e

        return result

    def _build_merged_cell_map(self, ws) -> Dict:
        """Build a map of merged cells → their top-left value."""
        merged_map = {}
        try:
            for merge_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merge_range.bounds
                value = ws.cell(min_row, min_col).value
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        # Skip the original cell
                        if row == min_row and col == min_col:
                            continue
                        merged_map[(row, col)] = value
        except Exception as e:
            logger.debug(f"Merged cell map building failed: {e}")
        return merged_map

    def _format_cell_value(self, val, cell=None) -> str:
        """Format a cell value preserving data type context."""
        if val is None:
            return ""

        # Handle dates
        if hasattr(val, 'strftime'):
            try:
                return val.strftime('%Y-%m-%d %H:%M:%S') if val.hour or val.minute else val.strftime('%Y-%m-%d')
            except Exception:
                return str(val)

        # Handle numbers — preserve precision
        if isinstance(val, float):
            if val == int(val):
                return str(int(val))
            return f"{val:.6g}"

        return str(val).strip()

    def _calculate_column_widths(self, rows_data: List[List[str]], max_width: int = 30) -> List[int]:
        """Calculate column widths for aligned text output."""
        if not rows_data:
            return []

        num_cols = max(len(row) for row in rows_data)
        widths = [0] * num_cols

        for row in rows_data:
            for i, cell in enumerate(row):
                if i < num_cols:
                    widths[i] = max(widths[i], min(len(cell), max_width))

        return [max(w, 3) for w in widths]  # Minimum width of 3

    def _extract_comments(self, ws) -> List[Dict[str, str]]:
        """Extract cell comments from a worksheet."""
        comments = []
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        comments.append({
                            "cell": cell.coordinate,
                            "author": cell.comment.author or "Unknown",
                            "text": cell.comment.text
                        })
        except Exception as e:
            logger.debug(f"Comment extraction failed: {e}")
        return comments

    def _extract_with_pandas(self, file_bytes: bytes, filename: str, file_id: str) -> TextExtractionResult:
        """Fallback extraction using pandas (simpler but no merged cell support)."""
        import pandas as pd

        result = TextExtractionResult(file_id, filename, "xlsx")
        result.extraction_method = "pandas"

        excel_file = pd.ExcelFile(io.BytesIO(file_bytes))

        for sheet_num, sheet_name in enumerate(excel_file.sheet_names, start=1):
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            sheet_text = f"Sheet: {sheet_name}\n\n{df.to_string(index=False)}"
            result.add_page(sheet_num, sheet_text, "pandas",
                            sheet_name=sheet_name,
                            rows=len(df),
                            columns=len(df.columns))

        result.metadata = {
            "sheet_count": len(excel_file.sheet_names),
            "sheet_names": excel_file.sheet_names
        }

        return result
